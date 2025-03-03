import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

# 비례대표선거 시도 code 조회, 파싱 함수
def get_city_codes():
    city_url = "http://info.nec.go.kr/bizcommon/selectbox/selectbox_cityCodeBySgJson.json"
    election_params = {
        'electionId': '0020240410',
        'electionCode': '7'
    }
    
    try:
        response = requests.get(city_url, params=election_params)
        data = response.json()
        return {city['CODE']: {
            'code': city['CODE'],
            'name': city['NAME'],
            'towns': get_town_codes(city['CODE'])
        } for city in data['jsonResult']['body']}
    except Exception as e:
        print(f"시도 코드 가져오기 실패: {e}")
        return None

# 비례대표선거 구시군 code 조회, 파싱 함수
def get_town_codes(city_code):
    town_url = "http://info.nec.go.kr/bizcommon/selectbox/selectbox_townCodeJson.json"
    election_params = {
        'electionId': '0020240410',
        'electionCode': '7',
        'cityCode': city_code
    }
    
    try:
        response = requests.get(town_url, params=election_params)
        data = response.json()
        return [{'code': town['CODE'], 'name': town['NAME']} 
                for town in data['jsonResult']['body']]
    except Exception as e:
        print(f"구시군 코드 가져오기 실패 (시도코드: {city_code}): {e}")
        return []

# 비례대표선거 데이터 조회, 파싱 함수
def get_election_data(city_code, town_code):
    url = "http://info.nec.go.kr/electioninfo/electionInfo_report.xhtml"
    params = {
        'electionId': '0020240410',
        'requestURI': '/electioninfo/0020240410/vc/vccp08.jsp',
        'topMenuId': 'VC',
        'secondMenuId': 'VCCP08',
        'menuId': 'VCCP08',
        'statementId': 'VCCP08_#7_1',
        'electionCode': '7',
        'cityCode': city_code,
        'sgaCityCode': '-1',
        'townCodeFromSgg': '-1',
        'townCode': town_code,
        'sggTownCode': '-1',
        'checkCityCode': '-1'
    }
    
    try:
        response = requests.get(url, params=params)
        soup = BeautifulSoup(response.text, 'html.parser')
        return parse_election_data(soup)
    except Exception as e:
        print(f"선거 데이터 가져오기 실패: {e}")
        return None

# HTML 파싱 - 개혁신당 관련 데이터 추출
def parse_election_data(soup):
    district_votes = {}
    rows = soup.select('table.table01 tbody tr')
    
    # 개혁신당 컬럼 인덱스 찾기
    reform_party_index = None
    header_row = soup.select('table.table01 thead tr')[1]
    headers = header_row.select('th')
    for idx, header in enumerate(headers):
        if '개혁신당' in header.text:
            reform_party_index = idx + 4
            break
    if reform_party_index is None:
        return None
        
    # 개혁신당 선거 데이터 추출
    for row in rows:
        columns = row.select('td')
        if len(columns) > 0:
            district_name = columns[0].text.strip()
            if district_name:
                try:
                    total_votes = int(columns[3].text.replace(',', ''))
                    reform_votes = int(columns[reform_party_index].text.replace(',', ''))
                    vote_rate = (reform_votes / total_votes * 100) if total_votes > 0 else 0
                    
                    district_votes[district_name] = {
                        '총 투표수': total_votes,
                        '개혁신당 득표수': reform_votes,
                        '개혁신당 득표율': round(vote_rate, 2)
                    }
                except (ValueError, IndexError) as e:
                    print(f"데이터 처리 중 오류 발생 ({district_name}): {e}")
    return district_votes

# 선거 결과를 엑셀 파일로 저장
def save_election_results(df, filename="2024비례대표선거_개혁신당_결과.xlsx"):
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        # 전체 데이터 시트 생성
        df.to_excel(writer, sheet_name='전체', index=False)
        
        # 시도별로 시트 생성
        for sido in df['시도'].unique():
            sido_df = df[df['시도'] == sido]
            sido_df.to_excel(writer, sheet_name=sido, index=False)
        
        # 각 시트에 스타일 적용
        for sheet_name in writer.sheets:
            apply_excel_styling(writer.sheets[sheet_name])
    
    print(f"\n전체 데이터가 {filename}에 저장되었습니다.")

# 엑셀 스타일 적용
def apply_excel_styling(worksheet):
    # 헤더 스타일
    header_fill = PatternFill(start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')
    header_font = Font(bold=True)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 헤더 스타일 적용
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # 데이터 영역 스타일 적용
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        for cell in row:
            cell.border = thin_border
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='center')
    
    # 구시군이 바뀌는 행에 굵은 선 추가
    prev_sigungu = None
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
        current_sigungu = row[1].value
        if prev_sigungu != current_sigungu:
            for cell in row:
                cell.border = Border(top=Side(style='medium'),
                                  left=Side(style='thin'),
                                  right=Side(style='thin'),
                                  bottom=Side(style='thin'))
        prev_sigungu = current_sigungu
    
    # 컬럼 너비 조정
    for col, width in zip('ABCDEF', [15, 15, 20, 15, 15, 15]):
        worksheet.column_dimensions[col].width = width

def main():
    # 시도, 구시군 code 조회
    city_codes = get_city_codes()
    all_votes = {}
    
    for city_code, city_info in city_codes.items():
        print(f"\n처리 중: {city_info['name']}")
        for town in city_info['towns']:
            print(f"  - {town['name']} 처리 중...")
            try:
                # 해당 시도, 구시군 선거 데이터 조회
                district_data = get_election_data(city_code, town['code'])
                if district_data:
                    for district, votes in district_data.items():
                        key = f"{city_info['name']}_{town['name']}_{district}"
                        all_votes[key] = {
                            '시도': city_info['name'],
                            '구시군': town['name'],
                            '읍면동': district,
                            **votes
                        }
            except Exception as e:
                print(f"  - {town['name']} 처리 중 오류 발생: {e}")
    
    df = pd.DataFrame.from_dict(all_votes, orient='index')
    df = df[['시도', '구시군', '읍면동', '총 투표수','개혁신당 득표수', '개혁신당 득표율']]
    save_election_results(df)

if __name__ == "__main__":
    main()